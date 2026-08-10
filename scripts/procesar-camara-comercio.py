# Extrae datos de Manrique (Comuna 3) de las tablas de estructura empresarial por comuna
# y genera un resumen en lib/data/camara-comercio-manrique.json para consumo estático del sitio.
#
# Tabla 14: empresas y activos por comuna y tamaño (Ley 905/2004)
# Tabla 15: empresas e ingresos por comuna y tamaño (Decreto 957/2019) — clasificación distinta
#           a la de activos, por eso NO se fusionan en una sola tabla "por tamaño".
# Tabla 18: empresas por sector (CIIU) y comuna — se toma el top de sectores en Manrique.
import json
import sys
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).parent.parent
XLSX = RAIZ / "data" / "camara-comercio" / "Estructura-Empresarial-2025.xlsx"
COMUNAS_VECINAS = ["ARANJUEZ", "SANTA CRUZ", "POPULAR", "VILLA HERMOSA"]

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)


def fila_comuna(nombre_tabla, comuna="MANRIQUE"):
    ws = wb[nombre_tabla]
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == comuna:
            return row
    raise ValueError(f'No encontré la fila "{comuna}" en {nombre_tabla}')


# --- Tabla 14: Comuna | Micro|Peq|Med|Grande | ActMicro|ActPeq|ActMed|ActGrande | TotalEmpresas | TotalActivos
t14 = fila_comuna("Tabla 14")
porTamanoActivos = [
    {"tamano": "Micro", "empresas": t14[1], "activos": t14[5]},
    {"tamano": "Pequeña", "empresas": t14[2], "activos": t14[6]},
    {"tamano": "Mediana", "empresas": t14[3], "activos": t14[7]},
    {"tamano": "Grande", "empresas": t14[4], "activos": t14[8]},
]

# --- Tabla 15: Comuna | Micro|Peq|Med|Grande|SinClasif | IngMicro|IngPeq|IngMed|IngGrande|IngSinClasif | TotalEmpresas | TotalIngresos
t15 = fila_comuna("Tabla 15")
porTamanoIngresos = [
    {"tamano": "Micro", "empresas": t15[1], "ingresos": t15[6]},
    {"tamano": "Pequeña", "empresas": t15[2], "ingresos": t15[7]},
    {"tamano": "Mediana", "empresas": t15[3], "ingresos": t15[8]},
    {"tamano": "Grande", "empresas": t15[4], "ingresos": t15[9]},
    {"tamano": "Sin clasificación", "empresas": t15[5], "ingresos": t15[10]},
]

# --- Tabla 14 completa: comparación de Manrique contra comunas vecinas (nororientales)
ws14 = wb["Tabla 14"]
comparacionComunas = []
for row in ws14.iter_rows(values_only=True):
    if row and (row[0] == "MANRIQUE" or row[0] in COMUNAS_VECINAS):
        comparacionComunas.append({"comuna": row[0].title(), "empresas": row[9], "activos": row[-1]})
orden = {"Manrique": 0, **{c.title(): i + 1 for i, c in enumerate(COMUNAS_VECINAS)}}
comparacionComunas.sort(key=lambda r: orden.get(r["comuna"], 99))

# --- Tabla 18: sectores (CIIU) con más empresas en Manrique
ws18 = wb["Tabla 18"]
header = next(ws18.iter_rows(min_row=4, max_row=4, values_only=True))
idx_manrique = header.index("MANRIQUE")

sectores = []
for row in ws18.iter_rows(min_row=5, values_only=True):
    if not row or row[0] is None or row[1] is None:
        continue
    cantidad = row[idx_manrique]
    if isinstance(cantidad, (int, float)) and cantidad > 0:
        sectores.append({"ciiu": row[0], "descripcion": row[1], "empresas": int(cantidad)})
sectores.sort(key=lambda r: r["empresas"], reverse=True)
sectores = sectores[:15]

resultado = {
    "fuente": "Cámara de Comercio de Medellín para Antioquia — Registro Mercantil",
    "cobertura": "Comuna 3 - Manrique",
    "periodo": "2025",
    "fechaProceso": "2026-07-21",
    "totalEmpresas": t14[9],
    "totalActivos": t14[-1],
    "totalIngresos": t15[-1],
    "porTamanoActivos": porTamanoActivos,
    "porTamanoIngresos": porTamanoIngresos,
    "comparacionComunas": comparacionComunas,
    "sectores": sectores,
    "metodologia": (
        "Comerciantes renovados y matriculados en el Registro Mercantil de la Cámara de "
        "Comercio de Medellín para Antioquia en 2025, georreferenciados en la Comuna 3 - "
        "Manrique. Tamaño por activos según Ley 905 de 2004 (Tabla 14); tamaño por ingresos "
        "según Decreto 957 de 2019 (Tabla 15) — son criterios distintos, por eso no coinciden "
        "exactamente las cantidades de empresas entre ambas clasificaciones. Sectores según "
        "clasificación CIIU (Tabla 18). Comunas vecinas elegidas por cercanía geográfica en la "
        "zona nororiental (Tabla 14)."
    ),
}

salida = RAIZ / "lib" / "data" / "camara-comercio-manrique.json"
salida.parent.mkdir(parents=True, exist_ok=True)
salida.write_text(json.dumps(resultado, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

sys.stdout.reconfigure(encoding="utf-8")
print(json.dumps(resultado, indent=2, ensure_ascii=False))
